from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from products.models import Product
from .forms import LoaderForm, BrandForm
from account.models import Brand
from .models import Loader, Matching, Rule
from products.models import Product
import openpyxl
from django.contrib import messages

@login_required(login_url='/login/')
def index(request):
    products = Product.objects.all()

    return render(request, 'account/index.html', {'products': products})


@login_required(login_url='/login/')
def addbrand(request):
    if request.method == 'POST':
        form = BrandForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('/profile/brands/')
            except:
                form.add_error(None, "Error with saving form")
    else:
        form = BrandForm
        return render(request, 'brand/add.html', {'form' : form})



@login_required(login_url='/login/')
def brands(request):
    brands = Brand.objects.all()
    return render(request, 'brand/index.html', {'brands': brands})


@login_required(login_url='/login/')
def showBrand(request, id):
    brand = Brand.objects.get(pk=id)
    loaders = Loader.objects.filter(brand_id=id)
    print(brand.name)
    return render(request, 'brand/show.html', {'brand': brand, 'loaders': loaders})




@login_required(login_url='/login/')
def show_loader(request, id, sheet_id=None):
    loader = Loader.objects.get(pk=id)
    book = openpyxl.open(loader.price, read_only=True)
    tabs = book.sheetnames

    # Find active page in Exl file
    if sheet_id is None:
        a = book.active
        print(a.title)
        j = 0
        for tab in tabs:
            if tab == a.title:
                active = j
            j = j + 1

        if active is None:
            sheet_id = 0
        else:
            sheet_id = int(active)

    # Load active page
    sheet = book[tabs[sheet_id]]

    # Get titles (first row) of active Excell page
    titles = []
    for col in range(0, sheet.max_column):
        titles.append(sheet[1][col].value)

    # Get Names of database fields
    db_titles = getColumnNames()


    matching = Matching.objects.filter(brand_id=loader.brand_id, sheet=sheet_id)
    try:
        rule = Rule.objects.get(brand_id=loader.brand_id, sheet=sheet_id)

    except:
        rule = None

    if len(matching) > 0:

        i = 0
        a = []
        for m in matching:
            #print(i)


            r = {
                'id': m.id,
                'db_index': m.db_index,
                'file_index': m.file_index,
                'brand_id': m.brand_id,
                'title': titles[i],
                'iteration': i
            }
            i += 1
            a.append(r)

    else:
        a = None

    if rule is not None:
        rule_id = int(rule.field_not_empty)
    else:
        rule_id = None

    return render(request, 'loader/show.html', {
        'loader': loader,
        'titles': titles,
        'db_titles': db_titles,
        'matching': a,
        'tabs': tabs,
        'sheet_id': sheet_id,
        'rule': rule,
        'rule_id': rule_id
    })





@login_required(login_url='/login/')
def save_matching(request, id):
    if request.method == 'POST':
        f = request.POST.getlist('f[]')
        d = request.POST.getlist('d[]')
        sheet_book = request.POST["sheet_book"]
        field_not_empty = request.POST["field_not_empty"]
        rules = request.POST["rules"]


        try:
            rule = Rule.objects.get(brand_id=id, sheet=sheet_book)
            return HttpResponse(id)
        except:
            rule = Rule.objects.create(brand_id=id, sheet=sheet_book)

        #Rule.objects.filter(brand_id=id, sheet=sheet_book).delete()
        Matching.objects.filter(brand_id=id, sheet=sheet_book).delete()


        rule.rules = rules
        rule.field_not_empty = field_not_empty
        rule.save()

        for i in f:
            i = int(i)
            m = Matching.objects.create(brand_id=id)
            m.db_index = d[i]
            m.file_index = i
            m.sheet = sheet_book

            m.rules = rules
            m.field_not_empty = field_not_empty

            m.save()

    messages.success(request, 'Changes successfully saved.')
    return redirect('/profile/loader/show/' + request.POST["loader_id"] + '/' + str(sheet_book) + '/', msg="Saved!")


@login_required(login_url='/login/')
def rem_loader(request, id, brand_id):
    Loader.objects.filter(pk=id).delete()
    messages.success(request, 'Deleted...')
    return redirect('/profile/brands/show/' + str(brand_id) + '/')


def copy_settings(doc):
    try:
        brand = Brand.objects.get(pk=doc.brand_id)
        rules = Rule.objects.filter(brand_id=brand.id)
        for r in rules:
            r.loader_id = doc.id
            r.save()
        return True
    except:
        return True

@login_required(login_url='/login/')
def add_loader(request):
    if request.method == 'POST':
        doc = Loader.objects.create()

        if request.FILES["price"]:
            doc.price =request.FILES["price"]

        #doc.user = request.user.username
        print (request.user.username)
        doc.brand_id = request.POST.get("brand_id")
        doc.user = request.user.username
        doc.save()

        copy_settings(doc)


        return redirect('/profile/brands/show/' + doc.brand_id + '/')
    else:
        form = LoaderForm
        return render(request, 'loader/add.html', {'form': form})




@login_required(login_url='/login/')
def run_file(request, id, sheet_id):
    z = 0
    loader = Loader.objects.get(pk=id)
    brand = Brand.objects.get(pk=loader.brand_id)

    try:
        rule = Rule.objects.get(brand_id=id, sheet=sheet_id)
    except:
        rule = None



    book = openpyxl.open(loader.price, read_only=True)
    tabs = book.sheetnames
    sheet = book[tabs[sheet_id]]

    matchings = Matching.objects.filter(brand_id=loader.brand_id, sheet=sheet_id)

##
    f = None
    w = None
    for mm in matchings:
        f = mm.field_not_empty
        wr = mm.rules
    f = int(f)
    #words = wr.split(",")

##


    db_titles = getColumnNames()


############


    for row in range(1, sheet.max_row+1):

        save = 1
        row = int(row)

        if str(sheet[row][f].value) == 'None':
            continue

        if str(sheet[row][f].value) == wr:
            print(wr + " " + str(sheet[row][f].value))
            continue

        product = Product.objects.create(brand=brand.name)
        product.key = rule.pk


        # i сдвиг вправо по колонкам
        #print(str(sheet[row][2].value))

        for col in range(0, sheet.max_column):
            #print (sheet[row][col].value)
            # sheet[1][col].value ## Title ROW
            for m in matchings:
                field = m.field_not_empty
                if m.file_index == col:
                    if db_titles[m.db_index] == 'None':
                        continue
                    else:
                        #product.__setattr__(db_titles[m.db_index], check_field(db_titles[m.db_index], sheet[row][col].value))
                        product.__setattr__(db_titles[m.db_index],sheet[row][col].value)
                        #print(sheet[row][col].value)

        if save is not None:
            product.save()
            z = z+1
            print(str(z) + " - " +str(product.pk))

    return HttpResponse("<h2> " + str(z) + "products saved...</h2> <br> <a href='/profile/products/'>Go to products</a>")


def rem_products(request, id=None):

    if id is not None:
        brand = Brand.objects.get(pk=id)
        Product.objects.filter(brand=brand.name).delete()
    else:
        Product.objects.filter().delete()

    return HttpResponse('Done')

# Additional Rules
def check_field(field, value):
    if field == 'msrp':
        value = str(value)
        value = value.replace("$", "")
        value = value.replace(" ", "")
        v = str(value)
        if len(v) > 6:
            v = float(v)
            v = round(v)
            return v
        return value

    if field == 'our_price':
        value = str(value)
        value = value.replace("$", "")
        value = value.replace(" ", "")
        v = str(value)
        if len(v) > 6:
            v = float(v)
            v = round(v)
            return v
        return value

    if field == 'map':
        value = str(value)
        value = value.replace("$", "")
        value = value.replace(" ", "")
        v = str(value)
        if len(v) > 6:
            v = float(v)
            v = round(v)
            return v
        return value

    if field == 'sku':
        v = str(value)
        return v[0:190]

    return value






def getColumnNames():
    arr = []
    i = 0
    for f in Product._meta.get_fields():
        if i == 0:
            arr.append('None')
            i = i + 1
            continue

        if str(f) == 'id':
            i = i + 1
            continue

        f = str(f)
        f = f[f.find(".") + 1 : ]
        arr.append(f[f.find(".") + 1 : ])
        i = i + 1
    return arr






@login_required(login_url='/login/')
def rules(request, loader_id):


    loader = Loader.objects.get(pk=loader_id)
    brand = Brand.objects.get(pk=loader.brand_id)

    return render(request, 'rules/index.html', {
        'brand': brand,
        'loader': loader,
    })


@login_required(login_url='/login/')
def add_rule(request, brand_id, loader_id=0, type=0):

    if type == 0:
        type='brand'


    sheet_id = 0
    brand = Brand.objects.get(pk=brand_id)

    try:
        loader = Loader.objects.get(pk=loader_id)
    except:
        loader = None

    row_num = 1



    try:
        rule = Rule.objects.get(brand_id=brand_id, loader_id=loader_id, type=type)
        row_num = int(rule.start_from)
        arr = rule.sheets.split(',')
        sheet_id = int(arr[0])
    except:
        try:
            rule = Rule.objects.get(brand_id=brand_id, loader_id=loader_id, type='brand')
            row_num = int(rule.start_from)
            arr = rule.sheets.split(',')
            sheet_id = int(arr[0])
        except:
            rule = None


    book = openpyxl.open(loader.price, read_only=True)
    tabs = book.sheetnames

    # Find active page in Exl file
    if sheet_id is None:
        a = book.active
        print(a.title)
        j = 0
        for tab in tabs:
            if tab == a.title:
                active = j
            j = j + 1

        if active is None:
            sheet_id = 0
        else:
            sheet_id = int(active)

    # Load active page
    sheet = book[tabs[sheet_id]]

    db_titles = [
        'None', 'sku', 'msrp', 'our_price', 'map'
    ]

    # Get titles (first row) of active Excell page
    titles = []
    for col in range(0, sheet.max_column):
        titles.append(sheet[row_num][col].value)


    print(titles)
    return render(request, 'rules/add.html', {
        'brand': brand,
        'loader': loader,
        'price_titles': titles,
        'tabs': tabs,
        'db_titles': db_titles,
        'type': type,
        'rule': rule
    })



@login_required(login_url='/login/')
def save_rule(request):
    if request.method == 'POST':
        print(request.POST)


        sheets = request.POST.getlist('sheets')
        sheets = ','.join(map(str, sheets))


        try:
            rule = Rule.objects.get(
                brand_id=int(request.POST.get('brand_id')),
                loader_id=int(request.POST.get('loader_id')),
                type=request.POST.get('type')
            )
        except:
            rule = Rule.objects.create(
                brand_id=int(request.POST.get('brand_id')),
                loader_id=int(request.POST.get('loader_id')),
                type=request.POST.get('type')
            )


        #return HttpResponse( makeInt(request.POST['start_from']) )
        rule.start_from = int(request.POST.get('start_from'))
        rule.sku = int(request.POST.get('sku'))
        rule.msrp = int(request.POST.get('msrp'))
        rule.our_price = int(request.POST.get('our_price'))
        rule.map = int(request.POST.get('map'))
        rule.sheets = sheets

        rule.rules = request.POST.get('rules')
        rule.field_not_empty = request.POST.get('field_not_empty')

        rule.remove_charter_all = request.POST.get('remove_charter_from_all_sku') #remove_charter_from_all_sku


        rule.remove_charter = request.POST.get('remove_charter_from_sku')
        rule.remove_charter_sku = request.POST.get('sku_to_remove_charter')

        rule.ignore_skus = request.POST.get('ignore_sku')

        rule.r_what = request.POST.get('what')
        rule.r_forwhat = request.POST.get('forwhat')

        rule.save()



        messages.success(request, 'Changes successfully saved.')
        return redirect('/profile/add-rule/' + str(rule.brand_id) + '/' + str(rule.loader_id) + '/' + str(rule.type) + '/', msg="Saved!")

def makeInt(s):
    s = str(s)
    s = s.replace(",", "")
    s = s.replace("(", "")
    s = s.replace(")", "")
    return int(s)



@login_required(login_url='/login/')
def run_file_new(request, rule_id):

    rule = Rule.objects.get(pk=rule_id)

    loader = Loader.objects.get(pk=rule.loader_id)
    brand = Brand.objects.get(pk=rule.brand_id)

    row_num = rule.start_from

    s = str(rule.sheets)

    if "," in s:
        s = s.split(',')

        all_prods = 0

        for tab in s:
            tab = int(tab)
            sheet_id = int(s[tab])

            z = load_proccess(sheet_id, loader, row_num, brand, rule)

            all_prods = all_prods + int(z)
    else:
        all_prods = 0
        s = int(s)
        sheet_id = int(s)

        z = load_proccess(sheet_id, loader, row_num, brand, rule)

        all_prods = all_prods + int(z)




    return HttpResponse("<h2> " + str(all_prods) + " products saved...</h2> <br> <a href='/profile/products/'>Go to products</a>")

def load_proccess(sheet_id, loader, row_num, brand, rule):

    book = openpyxl.open(loader.price, read_only=True)
    tabs = book.sheetnames
    sheet = book[tabs[sheet_id]]

    titles = ['None']
    for col in range(0, sheet.max_column):
        titles.append(sheet[row_num][col].value)

    z = 0

    for row in range(row_num, sheet.max_row + 1):

        save = 1
        row = int(row)
        # sheet[row][col].value

        product = Product()

        product.brand = brand.name

        for col in range(0, sheet.max_column):

            # SKU
            if int(col) == int(rule.sku):
                if rule.field_not_empty == 'sku':
                    if sheet[row][col].value == None or sheet[row][col].value == 'None' or sheet[row][
                        col].value == '' or str(sheet[row][col].value) == str(rule.rules):
                        save = 0

                sku = sheet[row][col].value
                if sku is not None:
                    sku = str(sku)
                    if rule.remove_charter_all is not None:
                        sku = sku.replace(rule.remove_charter_all, "")

                    if rule.remove_charter is not None:
                        if sku == str(rule.remove_charter_sku):
                            sku = sku.replace(rule.remove_charter, "")

                    if sku == rule.ignore_skus:
                        save = 0

                    if rule.ignore_skus is not None:
                        ign = str(rule.ignore_skus)
                        ign = ign.split(",")
                        for ig in ign:
                            if ig == sku:
                                save = 0

                    if rule.r_what is not None:
                        if rule.r_forwhat is not None:
                            sku = sku.replace(rule.r_what, rule.r_forwhat)

                product.sku = check_field('sku', sku)

            if int(col) == int(rule.msrp):
                if rule.field_not_empty == 'msrp':
                    if sheet[row][col].value == None or sheet[row][col].value == 'None' or sheet[row][
                        col].value == '' or str(sheet[row][col].value) == str(rule.rules):
                        save = 0

                value = str(sheet[row][col].value)
                value = value.replace("$", "")
                value = value.replace(" ", "")
                product.msrp = sheet[row][col].value

            if int(col) == int(rule.our_price):
                if rule.field_not_empty == 'our_price':
                    if sheet[row][col].value == None or sheet[row][col].value == 'None' or sheet[row][
                        col].value == '' or str(sheet[row][col].value) == str(rule.rules):
                        save = 0
                product.our_price = sheet[row][col].value

            if int(col) == int(rule.map):
                if rule.field_not_empty == 'map':
                    if sheet[row][col].value == None or sheet[row][col].value == 'None' or sheet[row][
                        col].value == '' or str(sheet[row][col].value) == str(rule.rules):
                        save = 0
                product.map = sheet[row][col].value

        if save == 1:
            z = z + 1
            product.save()
        else:
            del product

    return z