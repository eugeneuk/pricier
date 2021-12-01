from django.shortcuts import render, redirect
from django.http import HttpResponse
import openpyxl
import os

def index(request):

    dir = os.path.abspath(os.curdir)
    book = openpyxl.open(dir + "/pricer/book.xlsx", read_only=True)
    tab = book.sheetnames
    print(tab)
    #sheet = book["letter 2"]

    print(len(tab))
    sheet = book[tab[1]]


    print(sheet["B2"].value)
    #print(sheet[1][1].value)
    # Ряды с 1, ячейки с 0

    # for row in range(1, sheet.max_row+1):
    #     author = sheet[row][0].value
    #     title = sheet[row][1].value
    #     year = sheet[row][2].value
    #     print(author, " ", title, " ", year, sheet.max_row)
    #cells = sheet['B1':'C11']

    return HttpResponse("Hello!")

def show(request):
    return HttpResponse('''
    
    <h1>Hello World!!! It;s freeeee</h1>
    
    ''')